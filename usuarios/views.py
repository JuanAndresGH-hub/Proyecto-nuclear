from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_protect, ensure_csrf_cookie
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from .models import Usuario, Estudiante, DocenteAsesor, Instructor
from django.db import IntegrityError
from practicas.models import Practica, Evaluacion, ProyectoPractica
from django.utils import timezone
from django.db.models import Count, Q
import io
from datetime import datetime, timedelta


@ensure_csrf_cookie
@csrf_protect
def login_view(request):
    """Vista de inicio de sesión"""
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()

        # Log detallado
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"LOGIN ATTEMPT - Username: '{username}' (len={len(username)}), Password len: {len(password)}")

        # Debug: verificar que los datos se reciben correctamente
        print(f"\n=== LOGIN DEBUG ===", flush=True)
        print(f"POST data keys: {list(request.POST.keys())}", flush=True)
        print(f"Username from POST: '{username}'", flush=True)
        print(f"Password from POST: '{password}'", flush=True)
        print(f"Username length: {len(username)}", flush=True)
        print(f"Password length: {len(password)}", flush=True)

        # Intentar autenticar
        user = authenticate(request, username=username, password=password)
        print(f"Auth result: {user}", flush=True)
        print(f"=== END DEBUG ===\n", flush=True)

        if user is not None:
            login(request, user)
            # Redirigir según el rol
            if user.rol == 'COORDINADOR':
                return redirect('dashboard_coordinador')
            elif user.rol == 'ESTUDIANTE':
                return redirect('dashboard_estudiante')
            elif user.rol == 'DOCENTE':
                return redirect('dashboard_docente')
            elif user.rol == 'INSTRUCTOR':
                return redirect('dashboard_instructor')
        else:
            return render(request, 'login.html', {'error': 'Credenciales inválidas'})

    return render(request, 'login.html')


def logout_view(request):
    """Vista de cierre de sesión"""
    logout(request)
    return redirect('login')


@ensure_csrf_cookie
@csrf_protect
def register_view(request):
    """Vista de registro de nuevos usuarios"""
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        password_confirm = request.POST.get('password_confirm')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        rol = request.POST.get('rol', 'ESTUDIANTE')

        if password != password_confirm:
            return render(request, 'register.html', {'error': 'Las contraseñas no coinciden'})

        try:
            user = Usuario.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name,
                rol=rol
            )

            # Crear perfil según el rol
            if rol == 'ESTUDIANTE':
                Estudiante.objects.create(
                    usuario=user,
                    programa_academico='',
                    semestre=1
                )
            elif rol == 'DOCENTE':
                DocenteAsesor.objects.create(usuario=user)
            elif rol == 'INSTRUCTOR':
                Instructor.objects.create(usuario=user)

            login(request, user)
            return redirect('dashboard_estudiante')

        except IntegrityError:
            return render(request, 'register.html', {'error': 'El usuario ya existe'})

    return render(request, 'register.html')


@login_required
def dashboard_estudiante(request):
    """Dashboard para estudiantes"""
    context = {'user': request.user}

    if request.user.es_estudiante():
        context['perfil'] = request.user.perfil_estudiante

    return render(request, 'dashboard.html', context)


@login_required
def dashboard_docente(request):
    """Dashboard para docentes asesores"""
    from practicas.models import Practica

    context = {'user': request.user}

    if request.user.es_docente():
        perfil_docente = request.user.perfil_docente
        context['perfil'] = perfil_docente

        # Obtener las prácticas asignadas al docente
        practicas_asignadas = Practica.objects.filter(docente_asesor=perfil_docente).select_related('estudiante__usuario')

        # Obtener estudiantes únicos
        estudiantes_ids = practicas_asignadas.values_list('estudiante', flat=True).distinct()
        from usuarios.models import Estudiante
        estudiantes = Estudiante.objects.filter(id__in=estudiantes_ids).select_related('usuario')

        context['estudiantes_asignados'] = estudiantes
        context['total_estudiantes'] = estudiantes.count()
        context['practicas_activas'] = practicas_asignadas.filter(estado='EN_CURSO').count()

    return render(request, 'dashboard.html', context)


@login_required
def dashboard_instructor(request):
    """Dashboard para instructores"""
    from practicas.models import Practica

    context = {'user': request.user}

    if request.user.es_instructor():
        perfil_instructor = request.user.perfil_instructor
        context['perfil'] = perfil_instructor

        # Obtener las prácticas asignadas al instructor
        practicas_asignadas = Practica.objects.filter(instructor=perfil_instructor).select_related('estudiante__usuario')

        # Obtener practicantes únicos
        estudiantes_ids = practicas_asignadas.values_list('estudiante', flat=True).distinct()
        from usuarios.models import Estudiante
        practicantes = Estudiante.objects.filter(id__in=estudiantes_ids).select_related('usuario')

        context['practicantes'] = practicantes
        context['total_practicantes'] = practicantes.count()
        context['practicas_activas'] = practicas_asignadas.filter(estado='EN_CURSO').count()

    return render(request, 'dashboard.html', context)


@login_required
def dashboard_coordinador(request):
    """Dashboard para coordinadores"""
    if not request.user.es_coordinador():
        return JsonResponse({'error': 'No autorizado'}, status=403)

    from empresas.models import Empresa, Vacante, Postulacion

    # Estadísticas de estudiantes
    total_estudiantes = Estudiante.objects.count()
    estudiantes_con_hv_aprobada = Estudiante.objects.filter(estado_hv='APROBADA').count()
    estudiantes_con_hv_pendiente = Estudiante.objects.filter(estado_hv='PENDIENTE').count()

    # Estadísticas de prácticas
    total_practicas = Practica.objects.count()
    practicas_activas = Practica.objects.filter(estado='EN_CURSO').count()
    practicas_finalizadas = Practica.objects.filter(estado='FINALIZADA').count()

    # Estadísticas de empresas y vacantes
    total_empresas = Empresa.objects.filter(activa=True).count()
    total_vacantes = Vacante.objects.count()
    vacantes_activas = Vacante.objects.filter(estado='ACTIVA').count()

    # Estadísticas de postulaciones
    total_postulaciones = Postulacion.objects.count()
    postulaciones_pendientes = Postulacion.objects.filter(estado='EN_REVISION').count()
    postulaciones_aceptadas = Postulacion.objects.filter(estado='ACEPTADA').count()

    # Estadísticas de docentes e instructores
    total_docentes = DocenteAsesor.objects.count()
    total_instructores = Instructor.objects.count()

    # Hojas de vida pendientes (para el acceso rápido)
    hvs_pendientes = Estudiante.objects.filter(
        estado_hv='PENDIENTE'
    ).exclude(hv_archivo='').count()

    context = {
        'total_estudiantes': total_estudiantes,
        'estudiantes_con_hv_aprobada': estudiantes_con_hv_aprobada,
        'estudiantes_con_hv_pendiente': estudiantes_con_hv_pendiente,
        'hvs_pendientes': hvs_pendientes,
        'total_practicas': total_practicas,
        'practicas_activas': practicas_activas,
        'practicas_finalizadas': practicas_finalizadas,
        'total_empresas': total_empresas,
        'total_vacantes': total_vacantes,
        'vacantes_activas': vacantes_activas,
        'total_postulaciones': total_postulaciones,
        'postulaciones_pendientes': postulaciones_pendientes,
        'postulaciones_aceptadas': postulaciones_aceptadas,
        'total_docentes': total_docentes,
        'total_instructores': total_instructores,
    }

    return render(request, 'dashboard_coordinador.html', context)


@login_required
def profile_view(request):
    """Ver perfil del usuario"""
    context = {'user': request.user}

    if request.user.es_estudiante():
        context['perfil'] = request.user.perfil_estudiante
    elif request.user.es_docente():
        context['perfil'] = request.user.perfil_docente
    elif request.user.es_instructor():
        context['perfil'] = request.user.perfil_instructor

    return render(request, 'profile.html', context)


@login_required
@require_http_methods(["POST"])
def update_profile_view(request):
    """Actualizar perfil del usuario"""
    user = request.user

    user.first_name = request.POST.get('first_name', user.first_name)
    user.last_name = request.POST.get('last_name', user.last_name)
    user.email = request.POST.get('email', user.email)
    user.telefono = request.POST.get('telefono', user.telefono)
    user.save()

    return redirect('profile')


@login_required
def estudiantes_list(request):
    """Lista de estudiantes (solo para coordinador)"""
    if not request.user.es_coordinador():
        return JsonResponse({'error': 'No autorizado'}, status=403)

    estudiantes = Estudiante.objects.all()
    return render(request, 'estudiantes_list.html', {'estudiantes': estudiantes})


@login_required
def estudiante_detail(request, pk):
    """Detalle de un estudiante"""
    estudiante = get_object_or_404(Estudiante, pk=pk)

    # Solo el coordinador o el mismo estudiante pueden ver
    if not (request.user.es_coordinador() or request.user.id == estudiante.usuario.id):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    return render(request, 'estudiante_detail.html', {'estudiante': estudiante})


@login_required
@login_required
@csrf_protect
@require_http_methods(["POST"])
def upload_hoja_vida(request, pk):
    """Cargar hoja de vida"""
    import logging
    logger = logging.getLogger(__name__)

    logger.info(f"Upload HV - Request User: {request.user.username}, PK: {pk}")
    logger.info(f"Upload HV - FILES: {list(request.FILES.keys())}")
    logger.info(f"Upload HV - POST: {list(request.POST.keys())}")

    estudiante = get_object_or_404(Estudiante, pk=pk)

    # Solo el mismo estudiante puede subir su hoja de vida
    if request.user.id != estudiante.usuario.id:
        logger.error(f"Upload HV - No autorizado. User ID: {request.user.id}, Estudiante Usuario ID: {estudiante.usuario.id}")
        return JsonResponse({'error': 'No autorizado'}, status=403)

    if 'archivo' in request.FILES:
        logger.info(f"Upload HV - Archivo encontrado: {request.FILES['archivo'].name}")
        estudiante.hv_archivo = request.FILES['archivo']
        estudiante.estado_hv = 'PENDIENTE'
        estudiante.save()
        logger.info(f"Upload HV - Archivo guardado exitosamente para {estudiante.usuario.username}")
        return redirect('dashboard_estudiante')

    logger.error(f"Upload HV - No se proporcionó archivo")
    return JsonResponse({'error': 'No se proporcionó archivo'}, status=400)


@login_required
def docentes_list(request):
    """Lista de docentes asesores"""
    if not request.user.es_coordinador():
        return JsonResponse({'error': 'No autorizado'}, status=403)

    docentes = DocenteAsesor.objects.all()
    return render(request, 'docentes_list.html', {'docentes': docentes})


@login_required
def instructores_list(request):
    """Lista de instructores"""
    if not request.user.es_coordinador():
        return JsonResponse({'error': 'No autorizado'}, status=403)

    instructores = Instructor.objects.all()
    return render(request, 'instructores_list.html', {'instructores': instructores})


@login_required
def mis_practicas(request):
    """Vista de mis prácticas para estudiantes"""
    if not request.user.es_estudiante():
        return JsonResponse({'error': 'No autorizado'}, status=403)

    estudiante = request.user.perfil_estudiante
    practicas = Practica.objects.filter(estudiante=estudiante).order_by('numero')

    context = {
        'practicas': practicas,
        'total_practicas': practicas.count(),
        'practicas_aprobadas': practicas.filter(promedio_final__gte=3.5).count(),
    }
    return render(request, 'mis_practicas.html', context)


@login_required
def detalle_practica(request, pk):
    """Detalle de una práctica específica"""
    practica = get_object_or_404(Practica, pk=pk)

    # Verificar que el usuario es el dueño de la práctica o es docente/instructor asignado
    if not (request.user.id == practica.estudiante.usuario.id or
            request.user.id == (practica.docente_asesor.usuario.id if practica.docente_asesor else None) or
            request.user.id == (practica.instructor.usuario.id if practica.instructor else None)):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    evaluaciones = practica.evaluaciones.all()
    encuentros = practica.encuentros.all()
    # Como es OneToOneField, verificar si existe el proyecto
    proyecto = practica.proyecto if hasattr(practica, 'proyecto') else None

    context = {
        'practica': practica,
        'evaluaciones': evaluaciones,
        'encuentros': encuentros,
        'proyecto': proyecto,
        'promedio': practica.promedio_final,
    }
    return render(request, 'detalle_practica.html', context)


@login_required
def mis_documentos(request):
    """Vista de documentos del estudiante"""
    if not request.user.es_estudiante():
        return JsonResponse({'error': 'No autorizado'}, status=403)

    estudiante = request.user.perfil_estudiante

    # Obtener la hoja de vida
    hoja_vida = None
    if estudiante.hv_archivo:
        hoja_vida = {
            'tipo': 'Hoja de Vida',
            'nombre': estudiante.hv_archivo.name.split('/')[-1],
            'url': estudiante.hv_archivo.url,
            'estado': estudiante.estado_hv,
            'fecha': estudiante.fecha_creacion,
        }

    # Obtener proyectos (que contienen documentos)
    practicas = Practica.objects.filter(estudiante=estudiante)

    documentos = []
    for practica in practicas:
        # Como es OneToOneField, usar hasattr para verificar si existe
        if hasattr(practica, 'proyecto'):
            documentos.append({
                'proyecto': practica.proyecto,
                'practica': practica,
            })

    context = {
        'hoja_vida': hoja_vida,
        'documentos': documentos,
        'total_documentos': len(documentos) + (1 if hoja_vida else 0),
    }
    return render(request, 'mis_documentos.html', context)


@login_required
def ver_documentos_estudiante(request, estudiante_id):
    """Vista de documentos de un estudiante (para docentes/instructores)"""
    estudiante = get_object_or_404(Estudiante, pk=estudiante_id)

    # Verificar permisos: solo coordinador, docente asesor o instructor del estudiante
    tiene_permiso = False
    if request.user.es_coordinador():
        tiene_permiso = True
    elif request.user.es_docente():
        # Verificar si es docente de alguna práctica del estudiante
        tiene_permiso = Practica.objects.filter(
            estudiante=estudiante,
            docente_asesor=request.user.perfil_docente
        ).exists()
    elif request.user.es_instructor():
        # Verificar si es instructor de alguna práctica del estudiante
        tiene_permiso = Practica.objects.filter(
            estudiante=estudiante,
            instructor=request.user.perfil_instructor
        ).exists()

    if not tiene_permiso:
        return JsonResponse({'error': 'No autorizado'}, status=403)

    # Obtener la hoja de vida
    hoja_vida = None
    if estudiante.hv_archivo:
        hoja_vida = {
            'tipo': 'Hoja de Vida',
            'nombre': estudiante.hv_archivo.name.split('/')[-1],
            'url': estudiante.hv_archivo.url,
            'estado': estudiante.estado_hv,
            'fecha': estudiante.fecha_creacion,
        }

    # Obtener proyectos (que contienen documentos)
    practicas = Practica.objects.filter(estudiante=estudiante)

    documentos = []
    for practica in practicas:
        # Como es OneToOneField, usar hasattr para verificar si existe
        if hasattr(practica, 'proyecto'):
            documentos.append({
                'proyecto': practica.proyecto,
                'practica': practica,
            })

    context = {
        'estudiante': estudiante,
        'hoja_vida': hoja_vida,
        'documentos': documentos,
        'total_documentos': len(documentos) + (1 if hoja_vida else 0),
        'es_vista_docente': True,
    }
    return render(request, 'ver_documentos_estudiante.html', context)


@login_required
def mis_evaluaciones(request):
    """Vista de evaluaciones del estudiante"""
    if not request.user.es_estudiante():
        return JsonResponse({'error': 'No autorizado'}, status=403)

    estudiante = request.user.perfil_estudiante
    practicas = Practica.objects.filter(estudiante=estudiante)

    evaluaciones_por_practica = {}
    for practica in practicas:
        evaluaciones = practica.evaluaciones.all()
        if evaluaciones.exists():
            evaluaciones_por_practica[practica] = {
                'evaluaciones': evaluaciones,
                'promedio': practica.promedio_final,
                'estado': practica.estado,
            }

    context = {
        'evaluaciones_por_practica': evaluaciones_por_practica,
        'practicas': practicas,
    }
    return render(request, 'mis_evaluaciones.html', context)


@login_required
def revisar_hojas_vida(request):
    """Vista para que el coordinador revise hojas de vida pendientes"""
    if not request.user.es_coordinador():
        return JsonResponse({'error': 'No autorizado'}, status=403)

    # Obtener estudiantes con hojas de vida pendientes
    estudiantes_pendientes = Estudiante.objects.filter(
        estado_hv='PENDIENTE'
    ).exclude(hv_archivo='').select_related('usuario')

    # Obtener todas las hojas de vida para estadísticas
    total_aprobadas = Estudiante.objects.filter(estado_hv='APROBADA').count()
    total_rechazadas = Estudiante.objects.filter(estado_hv='RECHAZADA').count()
    total_pendientes = estudiantes_pendientes.count()

    context = {
        'estudiantes_pendientes': estudiantes_pendientes,
        'total_pendientes': total_pendientes,
        'total_aprobadas': total_aprobadas,
        'total_rechazadas': total_rechazadas,
    }

    return render(request, 'revisar_hojas_vida.html', context)


@login_required
@csrf_protect
@require_http_methods(["POST"])
def aprobar_rechazar_hoja_vida(request, estudiante_id):
    """Vista para aprobar o rechazar una hoja de vida"""
    if not request.user.es_coordinador():
        return JsonResponse({'error': 'No autorizado'}, status=403)

    estudiante = get_object_or_404(Estudiante, pk=estudiante_id)
    accion = request.POST.get('accion')  # 'aprobar' o 'rechazar'
    comentarios = request.POST.get('comentarios', '')

    if accion == 'aprobar':
        estudiante.estado_hv = 'APROBADA'
    elif accion == 'rechazar':
        estudiante.estado_hv = 'RECHAZADA'
    else:
        return JsonResponse({'error': 'Acción inválida'}, status=400)

    # Guardar comentarios si existen (necesitarías agregar este campo al modelo)
    estudiante.save()

    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"HV {accion.upper()} - Estudiante: {estudiante.usuario.username}, Por: {request.user.username}")

    return redirect('usuarios:revisar_hojas_vida')


@login_required
def reportes(request):
    """Vista de reportes y estadísticas para el coordinador"""
    if not request.user.es_coordinador():
        return JsonResponse({'error': 'No autorizado'}, status=403)

    from empresas.models import Empresa, Vacante, Postulacion

    # Estadísticas generales
    total_estudiantes = Estudiante.objects.count()
    total_practicas = Practica.objects.count()
    total_empresas = Empresa.objects.count()
    total_vacantes = Vacante.objects.count()
    total_docentes = DocenteAsesor.objects.count()
    total_instructores = Instructor.objects.count()

    # Estadísticas por estado
    estudiantes_con_hv = Estudiante.objects.exclude(hv_archivo='').count()
    practicas_activas = Practica.objects.filter(estado='EN_CURSO').count()
    practicas_finalizadas = Practica.objects.filter(estado='FINALIZADA').count()
    vacantes_activas = Vacante.objects.filter(estado='ACTIVA').count()

    # Postulaciones
    total_postulaciones = Postulacion.objects.count()
    postulaciones_aceptadas = Postulacion.objects.filter(estado='ACEPTADA').count()
    postulaciones_rechazadas = Postulacion.objects.filter(estado='RECHAZADA').count()
    postulaciones_pendientes = Postulacion.objects.filter(estado='EN_REVISION').count()

    # Prácticas por programa
    from django.db.models import Count
    practicas_por_programa = Practica.objects.values(
        'estudiante__programa_academico'
    ).annotate(total=Count('id')).order_by('-total')

    context = {
        'total_estudiantes': total_estudiantes,
        'total_practicas': total_practicas,
        'total_empresas': total_empresas,
        'total_vacantes': total_vacantes,
        'total_docentes': total_docentes,
        'total_instructores': total_instructores,
        'estudiantes_con_hv': estudiantes_con_hv,
        'practicas_activas': practicas_activas,
        'practicas_finalizadas': practicas_finalizadas,
        'vacantes_activas': vacantes_activas,
        'total_postulaciones': total_postulaciones,
        'postulaciones_aceptadas': postulaciones_aceptadas,
        'postulaciones_rechazadas': postulaciones_rechazadas,
        'postulaciones_pendientes': postulaciones_pendientes,
        'practicas_por_programa': practicas_por_programa,
    }

    return render(request, 'reportes.html', context)


@login_required
def exportar_estudiantes_excel(request):
    """Exportar listado completo de estudiantes a Excel"""
    if not request.user.es_coordinador():
        return JsonResponse({'error': 'No autorizado'}, status=403)

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return JsonResponse({'error': 'Librería openpyxl no instalada'}, status=500)

    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estudiantes"

    # Estilo del encabezado
    header_fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Encabezados
    headers = ['#', 'Código/Usuario', 'Nombre Completo', 'Email', 'Programa', 'Semestre',
               'HV Estado', 'Prácticas', 'Estado Práctica', 'Empresa Actual']

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Obtener estudiantes con prácticas
    estudiantes = Estudiante.objects.select_related('usuario').prefetch_related('practicas').all()

    # Llenar datos
    for idx, estudiante in enumerate(estudiantes, start=2):
        practica_actual = estudiante.practicas.filter(estado='EN_CURSO').first()
        total_practicas = estudiante.practicas.count()

        ws.cell(row=idx, column=1, value=idx-1)
        ws.cell(row=idx, column=2, value=estudiante.usuario.username)
        ws.cell(row=idx, column=3, value=estudiante.usuario.get_full_name())
        ws.cell(row=idx, column=4, value=estudiante.usuario.email)
        ws.cell(row=idx, column=5, value=estudiante.programa_academico)
        ws.cell(row=idx, column=6, value=estudiante.semestre)
        ws.cell(row=idx, column=7, value=estudiante.get_estado_hv_display())
        ws.cell(row=idx, column=8, value=total_practicas)

        if practica_actual:
            ws.cell(row=idx, column=9, value=practica_actual.get_estado_display())
            # Obtener empresa de la vacante si existe
            empresa = "Sin empresa"
            if hasattr(practica_actual, 'vacante') and practica_actual.vacante:
                empresa = practica_actual.vacante.empresa.nombre
            ws.cell(row=idx, column=10, value=empresa)
        else:
            ws.cell(row=idx, column=9, value="Sin práctica")
            ws.cell(row=idx, column=10, value="N/A")

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 25

    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=estudiantes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb.save(response)
    return response


@login_required
def exportar_practicas_pdf(request):
    """Exportar reporte de prácticas a PDF"""
    if not request.user.es_coordinador():
        return JsonResponse({'error': 'No autorizado'}, status=403)

    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    except ImportError:
        return JsonResponse({'error': 'Librería reportlab no instalada'}, status=500)

    # Crear buffer
    buffer = io.BytesIO()

    # Crear documento
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30,
                           topMargin=30, bottomMargin=30)

    # Contenedor de elementos
    elements = []

    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#198754'),
        spaceAfter=30,
        alignment=TA_CENTER
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#000000'),
        spaceAfter=12,
        spaceBefore=12
    )

    # Título
    title = Paragraph("Reporte de Prácticas Empresariales", title_style)
    elements.append(title)

    # Fecha
    fecha = Paragraph(f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal'])
    elements.append(fecha)
    elements.append(Spacer(1, 20))

    # Estadísticas generales
    from empresas.models import Empresa
    total_practicas = Practica.objects.count()
    practicas_activas = Practica.objects.filter(estado='EN_CURSO').count()
    practicas_finalizadas = Practica.objects.filter(estado='FINALIZADA').count()
    total_estudiantes = Estudiante.objects.count()

    stats_heading = Paragraph("Estadísticas Generales", heading_style)
    elements.append(stats_heading)

    stats_data = [
        ['Métrica', 'Cantidad'],
        ['Total de Prácticas', str(total_practicas)],
        ['Prácticas Activas', str(practicas_activas)],
        ['Prácticas Finalizadas', str(practicas_finalizadas)],
        ['Total de Estudiantes', str(total_estudiantes)],
    ]

    stats_table = Table(stats_data, colWidths=[4*inch, 2*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#198754')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(stats_table)
    elements.append(Spacer(1, 20))

    # Detalle de prácticas activas
    practicas_heading = Paragraph("Prácticas Activas", heading_style)
    elements.append(practicas_heading)

    practicas = Practica.objects.filter(estado='EN_CURSO').select_related(
        'estudiante__usuario', 'docente_asesor__usuario', 'instructor__usuario'
    )

    if practicas.exists():
        practicas_data = [['Estudiante', 'Programa', 'Docente', 'Instructor', 'Inicio']]

        for practica in practicas[:20]:  # Limitar a 20 para no hacer el PDF muy grande
            practicas_data.append([
                practica.estudiante.usuario.get_full_name(),
                practica.estudiante.programa_academico[:30],
                practica.docente_asesor.usuario.get_full_name() if practica.docente_asesor else 'N/A',
                practica.instructor.usuario.get_full_name() if practica.instructor else 'N/A',
                practica.fecha_inicio.strftime('%d/%m/%Y') if practica.fecha_inicio else 'N/A'
            ])

        practicas_table = Table(practicas_data, colWidths=[1.5*inch, 1.8*inch, 1.5*inch, 1.5*inch, 0.9*inch])
        practicas_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0dcaf0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black)
        ]))

        elements.append(practicas_table)
    else:
        elements.append(Paragraph("No hay prácticas activas", styles['Normal']))

    # Construir PDF
    doc.build(elements)

    # Preparar respuesta
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=reporte_practicas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'

    return response


@login_required
def estadisticas_avanzadas(request):
    """Vista de estadísticas avanzadas con gráficos"""
    if not request.user.es_coordinador():
        return JsonResponse({'error': 'No autorizado'}, status=403)

    from empresas.models import Empresa, Vacante, Postulacion

    # Datos para gráficos

    # 1. Prácticas por programa
    practicas_por_programa = Practica.objects.values(
        'estudiante__programa_academico'
    ).annotate(total=Count('id')).order_by('-total')

    # 2. Prácticas por estado
    practicas_por_estado = {
        'en_curso': Practica.objects.filter(estado='EN_CURSO').count(),
        'finalizada': Practica.objects.filter(estado='FINALIZADA').count(),
        'cancelada': Practica.objects.filter(estado='CANCELADA').count() if Practica.objects.filter(estado='CANCELADA').exists() else 0,
    }

    # 3. Estudiantes por estado de HV
    hv_estadisticas = {
        'pendiente': Estudiante.objects.filter(estado_hv='PENDIENTE').count(),
        'en_revision': Estudiante.objects.filter(estado_hv='EN_REVISION').count(),
        'aprobada': Estudiante.objects.filter(estado_hv='APROBADA').count(),
        'rechazada': Estudiante.objects.filter(estado_hv='RECHAZADA').count(),
    }

    # 4. Postulaciones por empresa (top 5)
    top_empresas = Postulacion.objects.values(
        'vacante__empresa__nombre'
    ).annotate(total=Count('id')).order_by('-total')[:5]

    # 5. Evolución de prácticas por mes (últimos 6 meses)
    from django.db.models.functions import TruncMonth
    practicas_por_mes = Practica.objects.filter(
        fecha_inicio__gte=timezone.now() - timedelta(days=180)
    ).annotate(
        mes=TruncMonth('fecha_inicio')
    ).values('mes').annotate(total=Count('id')).order_by('mes')

    # Convertir datos a JSON
    import json

    # Convertir QuerySet a lista de diccionarios simples
    practicas_por_mes_json = []
    for item in practicas_por_mes:
        practicas_por_mes_json.append({
            'mes': item['mes'].isoformat() if item['mes'] else None,
            'total': item['total']
        })

    context = {
        'practicas_por_programa': json.dumps(list(practicas_por_programa)),
        'practicas_por_estado': json.dumps(practicas_por_estado),
        'hv_estadisticas': json.dumps(hv_estadisticas),
        'top_empresas': json.dumps(list(top_empresas)),
        'practicas_por_mes': json.dumps(practicas_por_mes_json),
    }

    return render(request, 'estadisticas_avanzadas.html', context)

